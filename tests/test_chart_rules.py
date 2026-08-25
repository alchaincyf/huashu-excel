"""图表防误导规则的回归测试（无需 playwright / 浏览器，只依赖 openpyxl）。

覆盖 make_chart.py 落地的几条诚实性约束：
  · 饼图在类别 > PIE_MAX(3) 时拒绝生成
  · 柱/条形图数值轴强制从 0
  · 折线图不强制从 0
  · suggest() 对「双数值 + 分类」明确警告双 Y 轴
  · suggest() 对时间序列警告最后一期可能不完整

作者原帖强调这些规则是「有测量结果的，不是审美偏好」，
所以值得用可执行测试锁死，而不是只靠文档声称。

注：fixture 都用 ≥3 列——profile_table 对极端精简的 2 列表头检测会误判
（把单行表头当成两行），真实业务表（README 示例有 6 列）不受影响；
这里不测那个边界，只测图表规则本身。
"""
from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

REPO = Path(__file__).resolve().parents[1]
SCRIPTS = REPO / "scripts"
sys.path.insert(0, str(SCRIPTS))

import make_chart as mc  # noqa: E402


def _write_table(path: Path, header: list[str], rows: list[list]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for r in rows:
        ws.append(r)
    wb.save(path)


def test_pie_rejected_when_too_many_categories(tmp_path):
    """类别超过 PIE_MAX(3) 必须拒绝生成饼图，并给出 Cleveland-McGill 理由。"""
    p = tmp_path / "pie_bad.xlsx"
    _write_table(p, ["门店", "销售额", "区域"],
                 [[f"店{i}", 100 * (i + 1), f"华东{i % 2}"] for i in range(6)])  # 6 类
    out = tmp_path / "out.xlsx"
    with pytest.raises(SystemExit) as ei:
        mc.build(p, None, "门店", ["销售额"], "pie", "t", out)
    msg = str(ei.value)
    assert "饼图" in msg and "3 类" in msg, msg
    assert not out.exists(), "拒绝时不应生成文件"


def test_pie_allowed_within_limit(tmp_path):
    """类别 ≤ PIE_MAX 时饼图允许生成（回归：不能一刀切禁饼图）。"""
    p = tmp_path / "pie_ok.xlsx"
    _write_table(p, ["渠道", "占比", "区域"],
                 [["线上", 60, "国内"], ["线下", 40, "国内"]])  # 2 类
    out = tmp_path / "out.xlsx"
    mc.build(p, None, "渠道", ["占比"], "pie", "t", out)
    assert out.exists(), "≤3 类应成功生成饼图"


def test_bar_y_axis_forced_from_zero(tmp_path):
    """横向条形图：数值轴（x 轴）必须强制从 0。"""
    p = tmp_path / "bar.xlsx"
    _write_table(p, ["门店", "销售额", "区域"], [[f"店{i}", 100 + i, f"华东{i % 2}"] for i in range(4)])
    out = tmp_path / "out.xlsx"
    mc.build(p, None, "门店", ["销售额"], "bar", "t", out)
    wb = load_workbook(out)
    ch = wb["chart"]._charts[0]
    assert ch.x_axis.scaling.min == 0, "条形图数值轴必须从 0"


def test_col_y_axis_forced_from_zero(tmp_path):
    """竖向柱形图：Y 轴必须强制从 0。"""
    p = tmp_path / "col.xlsx"
    _write_table(p, ["门店", "销售额", "区域"], [[f"店{i}", 100 + i, f"华东{i % 2}"] for i in range(4)])
    out = tmp_path / "out.xlsx"
    mc.build(p, None, "门店", ["销售额"], "col", "t", out)
    wb = load_workbook(out)
    ch = wb["chart"]._charts[0]
    assert ch.y_axis.scaling.min == 0, "柱形图 Y 轴必须从 0"


def test_line_y_axis_not_forced_from_zero(tmp_path):
    """折线图：Y 轴不强制从 0（否则会压平真实波动）。"""
    p = tmp_path / "line.xlsx"
    _write_table(p, ["门店", "指数", "区域"], [[f"店{i}", 100 + i, f"华东{i % 2}"] for i in range(4)])
    out = tmp_path / "out.xlsx"
    mc.build(p, None, "门店", ["指数"], "line", "t", out)
    wb = load_workbook(out)
    ch = wb["chart"]._charts[0]
    assert ch.y_axis.scaling.min is None, "折线图不应强制 Y 轴从 0"


def test_suggest_warns_dual_axis(tmp_path):
    """双数值列 + 分类维度时，suggest 必须警告双 Y 轴风险。"""
    p = tmp_path / "dual.xlsx"
    _write_table(p, ["门店", "销售额", "利润"],
                 [[f"店{i}", 100 + i, 10 + i] for i in range(4)])
    prof = mc.profile(p, None)
    sugs = mc.suggest(prof)
    all_warns = [w for s in sugs for w in s.warnings]
    assert any("双 Y 轴" in w or "双轴" in w for w in all_warns), all_warns


def test_suggest_warns_incomplete_last_period(tmp_path):
    """时间序列（日期列 + 数值）时，suggest 必须警告最后一期可能不完整。"""
    p = tmp_path / "ts.xlsx"
    _write_table(p, ["月份", "销售额", "利润"],
                 [[f"2026-{m:02d}-01", 100 + m, 10 + m] for m in range(1, 13)])
    prof = mc.profile(p, None)
    sugs = mc.suggest(prof)
    all_warns = [w for s in sugs for w in s.warnings]
    assert any("最后一期" in w or "不完整" in w for w in all_warns), all_warns
