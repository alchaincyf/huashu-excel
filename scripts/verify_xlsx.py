#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Excel 交付物交付前的渲染自检——xlsx 版的闸门。

为什么需要它
------------
HTML 有 verify_visual.py、Word 有 verify_docx.py、数字有 verify_numbers.py，
唯独 xlsx 一直裸奔。于是这几类错只能靠人打开文件才发现，而它们
**在生成端全都不报错**——openpyxl 让你写什么就写什么：

  · 行高按「多少个字一行」估，没算字号 → 18 磅标题塞进 16 磅的行，上下被削
  · 数字列宽不够 → Excel 显示成 ####，一个数字都看不见
  · 长文本不合并也不换行，右边一格又有值 → 后半句被硬生生挡住
  · 合并区域盖住了已有数据 → 数据静默消失
  · 冻结窗格冻在空行上 → 真表头照样滚走，功能等于没开
  · 同级小节标题一半用强调色一半用墨色 → 层级表达自相矛盾

检查项（机器可判的事实，不判好看不好看）
  FAIL  行高装不下 wrap 后的文本
  FAIL  数值按 number_format 渲染后宽于列宽（会变 ####）
  FAIL  长文本靠溢出显示，但右邻单元格有值
  FAIL  合并区域内除左上角外还有非空单元格
  FAIL  冻结窗格冻的是空行
  WARN  浮点数没设 number_format
  WARN  长文本既没合并也没 wrap（右邻为空，只是溢出）
  WARN  数据表够长却没冻结表头
  WARN  同级标题（同字号同粗细）字体颜色不一致
  WARN  数据 sheet 裸奔：没列宽、没冻结、表头与正文无差别
  WARN  数值被显式设成左/居中对齐

用法
    python3 verify_xlsx.py 报告.xlsx
    python3 verify_xlsx.py 报告.xlsx --json
    python3 verify_xlsx.py --self-test          造几份已知有病的表，验闸门本身

退出码：有 FAIL 返回 1，否则 0。可直接当流水线闸门。

依赖：openpyxl + 标准库，没有别的。

一句话说明它的自我克制：**宽度类检查只在「这个 sheet 显式设过列宽」时才做。**
一列宽都没设的 sheet 是裸数据，不是排过版的交付物，拿版式标准去骂它
只会刷屏——那种情况由「数据 sheet 裸奔」一条 WARN 概括掉。
"""

from __future__ import annotations

import argparse
import json
import math
import os
import re
import sys
import tempfile
import unicodedata
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from xml.etree import ElementTree as ET

# ── 阈值：全部集中在这里，改这里就够了 ────────────────────────────
# 取值原则一律偏保守——宁可漏报，也不要用一堆边缘案例把真问题淹掉。
# verify_visual.py 的「数字接近但不相等」那一项就是反面教材：
# 误报率高到 workflow.md 里得专门写一段劝人忽略它。

BASE_FONT_SZ = 11.0        # 列宽单位的基准字号（Excel 默认 Calibri 11）
DEFAULT_FONT_SZ = 11.0     # cell.font.sz 缺省时按这个算
DEFAULT_COL_WIDTH = 8.43   # Excel 默认列宽（单位：基准字号下的字符宽）
DEFAULT_ROW_HEIGHT = 15.0  # Excel 默认行高（磅）

CJK_WIDTH = 2.0            # 全角字符宽度 = 2 个单位，ASCII = 1
# 单行占的磅数 ≈ 字号 × 1.35。references/xlsx-craft.md 的生成侧算法用 1.4
# 外加 2 磅内边距——那是刻意的：生成时宁可高一点（多几磅只是留白），
# 验收时宁可低一点（阈值松才不会误报）。两个数字站在同一个估算的两侧。
LINE_HEIGHT_RATIO = 1.35

ROW_HEIGHT_TOL = 1.08      # 需要的行高超出实际 8% 以内不报
ROW_SHORTFALL_MIN_PT = 2.5     # 且至少差这么多磅才报
ROW_SHORTFALL_MIN_FRAC = 0.25  # 且至少差 0.25 行才报（两个条件同时满足）

WIDTH_TOL = 1.05           # 文本/数字宽度超出列宽 5% 以内不报
LONG_TEXT_MIN_UNITS = 12.0     # 短标签溢出无所谓，宽度不到这个数不参与宽度检查

FLOAT_DECIMALS_MAX = 2     # 浮点数超过这么多位小数又没设格式 → 会露出二进制噪声
TABLE_MIN_ROWS = 30        # 超过这么多行的数据表才要求冻结表头
HEADING_MIN_SZ = 12.0      # 字号到这个数且加粗，才算「标题」，才查颜色一致性
HEADING_MIN_GROUP = 3      # 同级标题至少这么多个，才有「一致性」可言

COLLAPSE_AT = 3            # 同一 sheet 同一列同一类问题超过这么多条，折叠成一行
CROSS_SHEET_AT = 3         # 同一类问题横跨这么多个 sheet，再折一层
MAX_SCAN_ROWS = 20000      # 单 sheet 最多扫这么多行，防止百万行原始表把内存跑满

_FULLWIDTH = ("W", "F")

# 列宽的单位是「默认字体下最宽的那个数字有多宽」。把小写字母也按 1 算
# 会把英文文本的宽度高估三成——实测拿一份 Netflix 官方表跑，
# "Available Globally?"（19 个字符）在 20.8 宽的列里被误判成放不下。
_NARROW_CHARS = "iljtfrI.,;:'!|()[]{} -"
_LOWER_WIDTH = 0.75
_NARROW_WIDTH = 0.45


# ── 几何估算 ─────────────────────────────────────────────────────

def disp_width(text: str) -> float:
    """文本的显示宽度，单位 = 基准字号下的一个字符宽（≈ 一个数字的宽度）。
    全角按 2，数字与大写按 1，小写窄一些，i/l/. 这类更窄。"""
    w = 0.0
    for ch in text:
        if ch == "\t":
            w += 4.0
        elif unicodedata.east_asian_width(ch) in _FULLWIDTH:
            w += CJK_WIDTH
        elif ch in _NARROW_CHARS:
            w += _NARROW_WIDTH
        elif ch.islower():
            w += _LOWER_WIDTH
        else:
            w += 1.0
    return w


def capacity(col_width: float, font_sz: float) -> float:
    """一格能放下多少个「单位」。字号越大能放的越少。"""
    return col_width * (BASE_FONT_SZ / max(font_sz, 1.0))


def wrapped_lines(text: str, cap: float) -> int:
    """wrap_text 之后占几行。单元格里已有的显式换行符要分段各算各的。"""
    if cap <= 0:
        return 1
    n = 0
    for seg in str(text).split("\n"):
        n += max(1, math.ceil(disp_width(seg) / cap))
    return max(1, n)


# ── number_format 渲染估算 ───────────────────────────────────────
# 只要算出「渲染后有多少个字符宽」，不需要真的格式化。
# 覆盖常见写法：#,##0.## / 0.00 / 0.0% / "£"#,##0 / ¥#,##0.00 / General

_FMT_BRACKET = re.compile(r"\[[^\]]*\]")
_FMT_QUOTED = re.compile(r'"[^"]*"')


def render_len(value, fmt: str) -> float:
    """数值按 number_format 渲染后的显示宽度（单位同 disp_width）。"""
    if value is None:
        return 0.0
    sec = (fmt or "General").split(";")[0]

    if hasattr(value, "strftime"):
        # 日期时间：格式串里的 y/m/d/h/s 占位符数量约等于渲染长度
        body = _FMT_QUOTED.sub("", _FMT_BRACKET.sub("", sec))
        return float(len(body.replace("\\", "")) or 10)

    if sec.strip() in ("", "General", "@"):
        s = repr(float(value)) if isinstance(value, float) else str(value)
        if s.endswith(".0"):
            s = s[:-2]
        return disp_width(s)

    literals = "".join(m[1:-1] for m in _FMT_QUOTED.findall(sec))
    body = _FMT_QUOTED.sub("", _FMT_BRACKET.sub("", sec))
    pct = body.count("%")
    literals += "%" * pct
    # 反斜杠转义的字面量、以及货币符号这类直接写在格式里的字符
    literals += "".join(re.findall(r"\\(.)", body))
    body = re.sub(r"\\.", "", body)
    literals += "".join(ch for ch in body if ch in "¥$€£+-() ")

    num = float(value) * (100.0 ** pct)
    digits = body.replace(",", "").replace("%", "")
    int_part, _, dec_part = digits.partition(".")
    fixed = dec_part.count("0")
    optional = dec_part.count("#")

    frac = abs(num) - int(abs(num))
    need = 0
    if optional:
        s = f"{frac:.{fixed + optional}f}".rstrip("0")
        need = max(0, len(s) - 2)
    dec_n = max(fixed, min(need, fixed + optional))

    ip = f"{int(abs(num)):d}"
    ip = ip.rjust(int_part.count("0") or 1, "0")
    if "," in body:
        ip = f"{int(abs(num)):,}".rjust(len(ip) + (len(ip) - 1) // 3, "0")
    out = ip + ("." + "0" * dec_n if dec_n else "")
    if num < 0 and "-" not in literals:
        out = "-" + out
    return disp_width(out) + disp_width(literals)


# ── 读表 ────────────────────────────────────────────────────────

def color_key(color) -> str | None:
    """安全地把 openpyxl 的 Color 变成一个可比较的字符串。
    主题色的 .rgb 会返回一句描述符报错文本，不能直接拿来比。"""
    if color is None:
        return None
    for attr in ("rgb", "theme", "indexed"):
        v = getattr(color, attr, None)
        if isinstance(v, str) and not v.startswith("Values must be"):
            return f"{attr}:{v}"
        if isinstance(v, int):
            return f"{attr}:{v}"
    return None


def col_width_map(ws) -> tuple[dict[int, float], bool]:
    """列号 → 列宽。openpyxl 把 <col min=1 max=5 width=..> 这种成组声明
    只挂在起始列的字母下，直接按字母查会漏掉后面几列。"""
    widths: dict[int, float] = {}
    explicit = False
    for dim in ws.column_dimensions.values():
        if dim.width is None:
            continue
        explicit = True
        lo = dim.min or 1
        hi = dim.max or lo
        for i in range(lo, min(hi, ws.max_column) + 1):
            widths[i] = float(dim.width)
    return widths, explicit


def raw_nonempty(path: Path) -> dict[str, set[str]]:
    """直接读包里的 sheet XML，拿到「哪些坐标有值」。

    为什么不用 openpyxl：合并区域里非左上角的格子被 openpyxl 读成
    MergedCell，值一律是 None——正是要查的那份数据，在 API 层已经看不见了。
    只有翻原始 XML 才知道它当初写没写进去。"""
    out: dict[str, set[str]] = {}
    try:
        with zipfile.ZipFile(path) as z:
            names = set(z.namelist())
            if "xl/workbook.xml" not in names:
                return out
            ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
            rns = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
            wbx = ET.fromstring(z.read("xl/workbook.xml"))
            rels = {}
            if "xl/_rels/workbook.xml.rels" in names:
                for r in ET.fromstring(z.read("xl/_rels/workbook.xml.rels")):
                    rels[r.get("Id")] = r.get("Target")
            for sh in wbx.iter(f"{ns}sheet"):
                title = sh.get("name")
                tgt = rels.get(sh.get(f"{rns}id"))
                if not title or not tgt:
                    continue
                # Target 可能写成 worksheets/sheet1.xml、也可能写成
                # /xl/worksheets/sheet1.xml（openpyxl 写的就是后者）
                part = tgt.lstrip("/")
                if not part.startswith("xl/"):
                    part = "xl/" + part
                if part not in names:
                    continue
                coords: set[str] = set()
                sx = ET.fromstring(z.read(part))
                for c in sx.iter(f"{ns}c"):
                    if c.find(f"{ns}v") is not None or c.find(f"{ns}is") is not None:
                        ref = c.get("r")
                        if ref:
                            coords.add(ref)
                out[title] = coords
    except (zipfile.BadZipFile, ET.ParseError, KeyError):
        return out
    return out


# ── 问题记录 ────────────────────────────────────────────────────

class Issue:
    __slots__ = ("level", "kind", "sheet", "group", "coord", "msg", "sort")

    def __init__(self, level, kind, sheet, group, coord, msg, sort=0.0):
        self.level, self.kind, self.sheet = level, kind, sheet
        self.group, self.coord, self.msg, self.sort = group, coord, msg, sort


KIND_LABEL = {
    "row_height": "行高装不下 wrap 文本",
    "num_width": "数字列宽不足",
    "blocked": "长文本被右邻挡住",
    "overflow": "长文本靠溢出显示",
    "merge_eat": "合并吃掉了数据",
    "float_fmt": "浮点数没设 number_format",
    "freeze_blank": "冻结窗格冻在空行",
    "freeze_blank_col": "冻结窗格冻在空列",
    "no_freeze": "长数据表没冻结表头",
    "heading_color": "同级标题字色不一致",
    "naked": "数据 sheet 裸奔",
    "num_align": "数值被设成左/居中对齐",
}


def collapse(issues: list[Issue]) -> list[str]:
    """两层折叠，只为一件事：不刷屏。刷屏是这类脚本的头号敌人——
    verify_visual.py 有一项误报率高到 workflow.md 得专门劝人忽略它，
    结果是人开始整片跳过 WARN 区，连真问题一起漏掉。

      · 同一 sheet 同一列同类问题 > COLLAPSE_AT 条 → 折成一行
      · 同一类问题横跨 > CROSS_SHEET_AT 个 sheet → 再折成一行
        （12 个 sheet 的原始数据表，每个都「浮点没格式化」，说一次就够）
    """
    by_kind: dict[str, list[Issue]] = defaultdict(list)
    kind_order: list[str] = []
    for it in issues:
        if it.kind not in by_kind:
            kind_order.append(it.kind)
        by_kind[it.kind].append(it)

    lines: list[str] = []
    for kind in kind_order:
        pool = by_kind[kind]
        sheets = {i.sheet for i in pool}
        if len(sheets) > CROSS_SHEET_AT:
            worst = max(pool, key=lambda g: g.sort)
            lines.append(
                f"{KIND_LABEL.get(kind, kind)}：{len(sheets)} 个 sheet 共 {len(pool)} 处"
                f"（{'、'.join(sorted(sheets)[:3])} 等）。最严重的是 {worst.msg}")
            continue
        buckets: dict[tuple, list[Issue]] = defaultdict(list)
        order: list[tuple] = []
        for it in pool:
            k = (it.sheet, it.group)
            if k not in buckets:
                order.append(k)
            buckets[k].append(it)
        for k in order:
            grp = buckets[k]
            if len(grp) <= COLLAPSE_AT:
                lines.extend(g.msg for g in grp)
            else:
                worst = max(grp, key=lambda g: g.sort)
                lines.append(
                    f"「{k[0]}」{k[1]} 共 {len(grp)} 处同类问题，最严重的是 {worst.msg}")
    return lines


# ── 主检查 ──────────────────────────────────────────────────────

def check_sheet(ws, raw_coords: set[str], freeze_min_rows: int):
    fails: list[Issue] = []
    warns: list[Issue] = []
    notes: list[str] = []

    widths, has_widths = col_width_map(ws)
    dflt_w = float(ws.sheet_format.defaultColWidth or DEFAULT_COL_WIDTH)
    dflt_h = float(ws.sheet_format.defaultRowHeight or DEFAULT_ROW_HEIGHT)

    def cw(idx: int) -> float:
        return widths.get(idx, dflt_w)

    from openpyxl.utils import get_column_letter

    merged = list(ws.merged_cells.ranges)
    merge_of: dict[tuple[int, int], object] = {}
    for mr in merged:
        for rr in range(mr.min_row, mr.max_row + 1):
            for cc in range(mr.min_col, mr.max_col + 1):
                merge_of[(rr, cc)] = mr

    truncated = ws.max_row > MAX_SCAN_ROWS
    last_row = min(ws.max_row, MAX_SCAN_ROWS)
    if truncated:
        notes.append(f"「{ws.title}」共 {ws.max_row} 行，只扫了前 {MAX_SCAN_ROWS} 行")

    # 收集一遍，后面的检查都基于它
    cells = []
    for row in ws.iter_rows(min_row=1, max_row=last_row):
        for c in row:
            if c.value is not None:
                cells.append(c)

    sizes: Counter = Counter()
    colors: Counter = Counter()
    fills: set = set()
    headings: dict[tuple, list] = defaultdict(list)
    nonempty_per_row: dict[int, list[int]] = defaultdict(list)

    for c in cells:
        sz = float(c.font.sz or DEFAULT_FONT_SZ)
        sizes[sz] += 1
        ck = color_key(c.font.color)
        if ck:
            colors[ck] += 1
        if c.fill is not None and c.fill.patternType:
            fk = color_key(getattr(c.fill, "fgColor", None))
            if fk:
                fills.add(fk)
        nonempty_per_row[c.row].append(c.column)
        if isinstance(c.value, str) and c.font.b and sz >= HEADING_MIN_SZ:
            headings[(sz, True)].append(c)

    # ① 行高装不下 wrap 后的文本
    #    真实故障：make_report.py 按「46 个字一行 × 15 磅」估行高，没算字号，
    #    18 磅的大标题塞进 16 磅的行，上下各削掉一截。
    #    只在行高被显式设过时才判——没设的行 Excel 会自动撑开，不构成问题。
    if has_widths:
        for r in range(1, last_row + 1):
            dim = ws.row_dimensions.get(r) if hasattr(ws.row_dimensions, "get") else None
            h = getattr(dim, "height", None) if dim is not None else None
            if h is None:
                continue
            worst = None
            for c in ws[r] if r <= last_row else []:
                if c.value is None or not isinstance(c.value, str):
                    continue
                if not c.alignment.wrap_text:
                    continue
                mr = merge_of.get((c.row, c.column))
                if mr is not None:
                    if mr.min_row != mr.max_row:
                        continue           # 跨行合并，行高不能单独归给这一行
                    total_w = sum(cw(i) for i in range(mr.min_col, mr.max_col + 1))
                    if (c.row, c.column) != (mr.min_row, mr.min_col):
                        continue
                else:
                    total_w = cw(c.column)
                sz = float(c.font.sz or DEFAULT_FONT_SZ)
                line_h = sz * LINE_HEIGHT_RATIO
                n = wrapped_lines(c.value, capacity(total_w, sz))
                need = n * line_h
                if worst is None or need > worst[0]:
                    worst = (need, n, line_h, c)
            if worst is None:
                continue
            need, n, line_h, c = worst
            short = need - float(h)
            if (need > float(h) * ROW_HEIGHT_TOL
                    and short >= max(ROW_SHORTFALL_MIN_PT, ROW_SHORTFALL_MIN_FRAC * line_h)):
                sz = float(c.font.sz or DEFAULT_FONT_SZ)
                if n == 1:
                    why = (f"{sz:.0f} 磅的字至少要 {line_h:.1f} 磅行高，"
                           f"实际只有 {float(h):.0f} 磅 → 上下被削掉一截")
                else:
                    why = (f"文本 wrap 后需 {n} 行 × {line_h:.1f} 磅 = {need:.1f} 磅，"
                           f"实际行高 {float(h):.0f} 磅 → 底部约 "
                           f"{short / line_h:.1f} 行被裁")
                fails.append(Issue(
                    "FAIL", "row_height", ws.title, f"第 {r} 行", c.coordinate,
                    f"「{ws.title}」第 {r} 行（{c.coordinate}）：{why}。"
                    f"建议设为 {math.ceil(need)}",
                    sort=short))

    # ② 数字列宽不足 —— Excel 会整格显示成 ####，一个数字都看不见
    #    真实故障场景：千分位 + 货币符号 + 两位小数把 12 个字符塞进 10 宽的列。
    if has_widths:
        for c in cells:
            if not isinstance(c.value, (int, float)) or isinstance(c.value, bool):
                continue
            mr = merge_of.get((c.row, c.column))
            if mr is not None:
                if (c.row, c.column) != (mr.min_row, mr.min_col):
                    continue
                total_w = sum(cw(i) for i in range(mr.min_col, mr.max_col + 1))
            else:
                total_w = cw(c.column)
            if c.column not in widths:
                continue                     # 这一列没显式设宽，用户自己拉就行
            sz = float(c.font.sz or DEFAULT_FONT_SZ)
            need = render_len(c.value, c.number_format)
            cap = capacity(total_w, sz)
            if need > cap * WIDTH_TOL and need >= LONG_TEXT_MIN_UNITS * 0.5:
                letter = get_column_letter(c.column)
                fails.append(Issue(
                    "FAIL", "num_width", ws.title, f"{letter} 列", c.coordinate,
                    f"「{ws.title}」{c.coordinate}：{c.value!r} 按格式 "
                    f"{c.number_format!r} 渲染需 {need:.1f} 字宽，"
                    f"{letter} 列只有 {cap:.1f}（{sz:.0f} 磅字）→ Excel 显示成 ####。"
                    f"建议 {letter} 列宽 ≥ {math.ceil(need * sz / BASE_FONT_SZ) + 1}",
                    sort=need - cap))

    # ③/⑥ 长文本溢出：右边有值 = 被挡住（FAIL），右边为空 = 只是溢出（WARN）
    #    真实故障场景：三个 KPI 的说明文字并排放在 16 宽的列里，
    #    第一条「2,741 户，£7,136,227 → £6,553,605」直接被第二条压掉后半句。
    if has_widths:
        for c in cells:
            if not isinstance(c.value, str) or c.alignment.wrap_text:
                continue
            if (c.row, c.column) in merge_of:
                continue
            if c.column not in widths:
                continue
            sz = float(c.font.sz or DEFAULT_FONT_SZ)
            need = disp_width(c.value)
            cap = capacity(cw(c.column), sz)
            if need <= max(cap * WIDTH_TOL, LONG_TEXT_MIN_UNITS):
                continue
            letter = get_column_letter(c.column)
            nxt = ws.cell(row=c.row, column=c.column + 1)
            want = math.ceil(need * sz / BASE_FONT_SZ) + 1
            if nxt.value is not None:
                fails.append(Issue(
                    "FAIL", "blocked", ws.title, f"{letter} 列", c.coordinate,
                    f"「{ws.title}」{c.coordinate}：「{str(c.value)[:24]}」需 {need:.0f} 字宽，"
                    f"{letter} 列只有 {cap:.0f}，而右邻 {nxt.coordinate} 有值 → "
                    f"超出部分被挡掉。把 {letter} 列拉到 {want}、或合并到右边、"
                    f"或开 wrap_text 并配足行高",
                    sort=need - cap))
            elif len(nonempty_per_row.get(c.row, [])) >= 2:
                # 只在「这一行还有别的内容」时才提醒。整行只有它一个的
                # 标题/脚注行，靠溢出铺开是 Excel 里完全正常的写法，
                # 右边永远不会被填上——为它报警只会刷屏。
                warns.append(Issue(
                    "WARN", "overflow", ws.title, f"{letter} 列", c.coordinate,
                    f"「{ws.title}」{c.coordinate}：「{str(c.value)[:24]}」需 {need:.0f} 字宽，"
                    f"{letter} 列只有 {cap:.0f}，靠溢出到右边显示。"
                    f"现在右邻是空的所以看得见，一旦那格填上值就被挡住——"
                    f"要么合并，要么开 wrap_text",
                    sort=need - cap))

    # ④ 合并区域盖住了已有数据 —— 数据静默消失，且事后完全看不出来
    for mr in merged:
        anchor = f"{get_column_letter(mr.min_col)}{mr.min_row}"
        buried = sorted(
            f"{get_column_letter(cc)}{rr}"
            for rr in range(mr.min_row, mr.max_row + 1)
            for cc in range(mr.min_col, mr.max_col + 1)
            if f"{get_column_letter(cc)}{rr}" != anchor
            and f"{get_column_letter(cc)}{rr}" in raw_coords)
        if buried:
            fails.append(Issue(
                "FAIL", "merge_eat", ws.title, str(mr), anchor,
                f"「{ws.title}」合并区域 {mr} 里除左上角 {anchor} 外还有 "
                f"{len(buried)} 个非空单元格（{'、'.join(buried[:4])}"
                f"{' 等' if len(buried) > 4 else ''}）→ 合并会吃掉它们，数据静默丢失。"
                f"先把这些值挪走再合并，或者干脆不合并",
                sort=len(buried)))

    # ⑤ 浮点数没设 number_format —— 会原样显示成 9351342.370000001
    for c in cells:
        if not isinstance(c.value, float) or isinstance(c.value, bool):
            continue
        if c.number_format not in ("General", "", None):
            continue
        s = repr(c.value)
        dec = len(s.split(".")[1]) if "." in s else 0
        if dec > FLOAT_DECIMALS_MAX:
            warns.append(Issue(
                "WARN", "float_fmt", ws.title, "浮点格式", c.coordinate,
                f"「{ws.title}」{c.coordinate}：{s} 是浮点数且 number_format 为 General，"
                f"会连二进制噪声一起显示（{dec} 位小数）。"
                f"设成 '#,##0.00' 或 '0.0%' 之类的显式格式",
                sort=dec))

    # 表格形态判断：给⑦⑨用
    modal_cols = Counter(len(v) for v in nonempty_per_row.values())
    modal = modal_cols.most_common(1)[0] if modal_cols else (0, 0)
    looks_table = (len(nonempty_per_row) >= freeze_min_rows
                   and modal[0] >= 2
                   and modal[1] >= 0.6 * len(nonempty_per_row))

    first_row = min(nonempty_per_row) if nonempty_per_row else None
    header_styled = False
    if first_row is not None:
        def style_of(c):
            return (float(c.font.sz or DEFAULT_FONT_SZ), bool(c.font.b),
                    color_key(c.font.color))
        hdr = [style_of(c) for c in cells if c.row == first_row]
        body = Counter(style_of(c) for c in cells if c.row > first_row)
        if hdr and body:
            # 跟正文里最常见的那一种比。用集合差会被正文里零星的粗体小标题
            # 抵消掉，从而把「表头确实有样式」误判成没样式。
            modal_style = body.most_common(1)[0][0]
            header_styled = any(h != modal_style for h in hdr)

    fp = ws.freeze_panes

    # ⑨ 数据 sheet 裸奔 —— 先算，因为它一条就把「没冻结」概括进去了，
    #    两条一起报等于同一件事说两遍。
    naked_missing: list[str] = []
    if looks_table:
        if not has_widths:
            naked_missing.append("没设任何列宽")
        if not fp:
            naked_missing.append("没有冻结窗格")
        if not header_styled:
            naked_missing.append("表头与正文字体完全一样")
    if len(naked_missing) >= 2:
        warns.append(Issue(
            "WARN", "naked", ws.title, "整表", "",
            f"「{ws.title}」是 {len(nonempty_per_row)} 行的数据表，但"
            f"{'、'.join(naked_missing)}——交付物不该长这样。"
            f"（若它只是图表的数据源、不打算给人看，忽略这条）",
            sort=len(naked_missing)))

    # ⑦ 冻结窗格
    #    真实故障场景：内容从 B2 开始，freeze_panes 却写成 "A2"——
    #    冻住的是完全空白的第 1 行，真表头（第 2 行）照样滚走。
    if fp:
        m = re.match(r"^([A-Z]+)(\d+)$", str(fp))
        if m:
            from openpyxl.utils import column_index_from_string
            fcol = column_index_from_string(m.group(1))
            frow = int(m.group(2))
            rows_frozen = list(range(1, frow))
            cols_frozen = list(range(1, fcol))
            if rows_frozen and not any(r in nonempty_per_row for r in rows_frozen):
                fails.append(Issue(
                    "FAIL", "freeze_blank", ws.title, "冻结窗格", str(fp),
                    f"「{ws.title}」freeze_panes={fp}，冻住的第 "
                    f"{'、'.join(map(str, rows_frozen))} 行全是空的，"
                    f"而首个有内容的行是第 {first_row} 行 → 冻了一条空白带，"
                    f"真表头照样滚走。改成 '{m.group(1)}{(first_row or 1) + 1}'",
                    sort=len(rows_frozen)))
            if cols_frozen and not any(
                    any(cc in v for v in nonempty_per_row.values()) for cc in cols_frozen):
                warns.append(Issue(
                    "WARN", "freeze_blank_col", ws.title, "冻结窗格", str(fp),
                    f"「{ws.title}」freeze_panes={fp}，冻住的前 {fcol - 1} 列全是空的。"
                    f"如果那是刻意留的装订边可以忽略，否则冻结列没起作用",
                    sort=1))
    elif looks_table and has_widths and len(naked_missing) < 2:
        warns.append(Issue(
            "WARN", "no_freeze", ws.title, "冻结窗格", "",
            f"「{ws.title}」是 {len(nonempty_per_row)} 行 × {modal[0]} 列的数据表却没有冻结窗格，"
            f"往下翻就看不到列名了。设 ws.freeze_panes = "
            f"'A{(first_row or 1) + 1}'",
            sort=len(nonempty_per_row)))

    # ⑧ 同级标题字色不一致
    #    真实故障场景：7 个 13 磅粗体小节标题，只有前 2 个用了强调色，其余是墨色。
    for (sz, _b), group in sorted(headings.items()):
        if len(group) < HEADING_MIN_GROUP:
            continue
        by_color: dict[str | None, list] = defaultdict(list)
        for c in group:
            by_color[color_key(c.font.color)].append(c)
        if len(by_color) < 2:
            continue
        parts = []
        for ck, cs in sorted(by_color.items(), key=lambda kv: -len(kv[1])):
            sample = "、".join(c.coordinate for c in cs[:3])
            parts.append(f"{ck or '默认色'} × {len(cs)}（{sample}"
                         f"{' 等' if len(cs) > 3 else ''}）")
        warns.append(Issue(
            "WARN", "heading_color", ws.title, f"{sz:.0f} 磅粗体", "",
            f"「{ws.title}」{len(group)} 个 {sz:.0f} 磅粗体标题用了 {len(by_color)} 种字色："
            f"{'；'.join(parts)} → 同一层级表达不一致，读者会以为是两种层级。"
            f"统一成一种，或者把想强调的那几个改成不同字号",
            sort=len(by_color)))

    # ⑨ 数据 sheet 裸奔 —— 在上面第 ⑦ 项之前就报过了，这里不重复

    # ⑩ 数值被显式设成左/居中对齐
    #    只报显式设过的：openpyxl 里数值默认（horizontal=None）就是右对齐，
    #    把 None 也算进来会让每一张没排过版的表全线飘红。
    for c in cells:
        if not isinstance(c.value, (int, float)) or isinstance(c.value, bool):
            continue
        h = c.alignment.horizontal
        if h in ("left", "center"):
            warns.append(Issue(
                "WARN", "num_align", ws.title, "数值对齐", c.coordinate,
                f"「{ws.title}」{c.coordinate}：数值被显式设成 {h} 对齐，"
                f"同列数字的个位对不齐，一眼比不出大小。删掉 horizontal 让它默认右对齐",
                sort=1))

    info = {
        "sheet": ws.title,
        "rows": ws.max_row, "cols": ws.max_column,
        "merged": len(merged),
        "freeze": fp or None,
        "font_sizes": sorted(sizes),
        "font_colors": len(colors),
        "fills": len(fills),
        "explicit_widths": len(widths),
    }
    return fails, warns, notes, info


def verify(path: Path, freeze_min_rows: int):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=False)
    raw = raw_nonempty(path)
    fails: list[Issue] = []
    warns: list[Issue] = []
    notes: list[str] = []
    infos: list[dict] = []
    for ws in wb.worksheets:
        f, w, n, i = check_sheet(ws, raw.get(ws.title, set()), freeze_min_rows)
        fails += f
        warns += w
        notes += n
        infos.append(i)
    if not raw:
        notes.append("包内 sheet XML 没读成，「合并吃掉数据」这一项没验成——不等于通过")
    return fails, warns, notes, infos


# ── 自检：造几份「已知有病」的表，断言全被拦下 ──────────────────
# 闸门自己也会烂。改过这个文件之后跑一次 --self-test，
# 确认它还抓得住它当初就是为了抓的那几类错，并且不冤枉正确的写法。

def self_test() -> int:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font
    except ImportError:
        print("自检需要 openpyxl", file=sys.stderr)
        return 2

    tmp = tempfile.mkdtemp(prefix="verify_xlsx_selftest_")
    ok = True

    def build(fn):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "报告"
        fn(ws)
        p = os.path.join(tmp, f"{fn.__name__}.xlsx")
        wb.save(p)
        return Path(p)

    def case_row_height(ws):
        ws.column_dimensions["B"].width = 40
        c = ws.cell(row=2, column=2, value="标" * 60)
        c.font = Font(size=11)
        c.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[2].height = 16

    def case_num_width(ws):
        ws.column_dimensions["B"].width = 6
        c = ws.cell(row=2, column=2, value=9351342.37)
        c.number_format = "#,##0.00"

    def case_blocked(ws):
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 10
        ws.cell(row=2, column=2, value="这是一句被右边挡掉后半截的说明文字")
        ws.cell(row=2, column=3, value="挡路的")

    def case_freeze_blank(ws):
        ws.column_dimensions["B"].width = 20
        ws.cell(row=2, column=2, value="表头")
        ws.cell(row=3, column=2, value="数据")
        ws.freeze_panes = "A2"

    def case_heading_color(ws):
        ws.column_dimensions["B"].width = 40
        for i, col in enumerate(["FF17557A", "FF17557A", "FF16181D",
                                 "FF16181D", "FF16181D"]):
            ws.cell(row=2 + i * 2, column=2, value=f"小节 {i}").font = \
                Font(size=13, bold=True, color=col)

    def case_clean(ws):
        # 正确写法：列宽够、行高按字号配足、冻结在真表头下面、标题同色
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 14
        t = ws.cell(row=1, column=2, value="标题")
        t.font = Font(size=18, bold=True, color="FF17557A")
        t.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[1].height = 26
        b = ws.cell(row=2, column=2, value="正" * 30)
        b.font = Font(size=11)
        b.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[2].height = 32
        n = ws.cell(row=2, column=3, value=1234.5)
        n.number_format = "#,##0.00"
        ws.freeze_panes = "A3"

    def case_merge_eat(ws):
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20
        ws["A1"] = "标题"
        ws["B1"] = "会被合并吃掉的值"

    def patch_merge(p: Path) -> Path:
        """openpyxl 存盘时会把合并区里非左上角的值清掉，用它造不出这个病灶。
        直接往包里塞一条 mergeCell，模拟别的工具合并到已有数据上的情形。"""
        out = p.with_name("merge_eat_patched.xlsx")
        with zipfile.ZipFile(p) as zin, \
                zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
            for n in zin.namelist():
                d = zin.read(n)
                if n.endswith("worksheets/sheet1.xml"):
                    d = d.replace(
                        b"</sheetData>",
                        b'</sheetData><mergeCells count="1">'
                        b'<mergeCell ref="A1:B1"/></mergeCells>')
                zout.writestr(n, d)
        return out

    f, _, _, _ = verify(patch_merge(build(case_merge_eat)), TABLE_MIN_ROWS)
    hit = any(i.kind == "merge_eat" for i in f)
    print(f"  {'✓' if hit else '✗'} 合并区域吃掉了已有数据")
    ok = ok and hit

    cases = [
        ("行高装不下 wrap 文本", case_row_height, "row_height", "FAIL"),
        ("数字列宽不足会显示 ####", case_num_width, "num_width", "FAIL"),
        ("长文本被右邻挡住", case_blocked, "blocked", "FAIL"),
        ("冻结窗格冻在空行", case_freeze_blank, "freeze_blank", "FAIL"),
        ("同级标题字色不一致", case_heading_color, "heading_color", "WARN"),
    ]
    for name, fn, kind, level in cases:
        f, w, _, _ = verify(build(fn), TABLE_MIN_ROWS)
        pool = f if level == "FAIL" else w
        hit = any(i.kind == kind for i in pool)
        print(f"  {'✓' if hit else '✗'} {name}")
        ok = ok and hit

    f, w, _, _ = verify(build(case_clean), TABLE_MIN_ROWS)
    clean = not f and not w
    print(f"  {'✓' if clean else '✗'} 正确写法不误报"
          + ("" if clean else "：" + "；".join(i.msg for i in f + w)))
    ok = ok and clean

    print("\n自检" + ("通过" if ok else "未通过——闸门坏了，先修它再用它验别的表"))
    return 0 if ok else 1


def main() -> int:
    ap = argparse.ArgumentParser(description="Excel 交付物交付前自检")
    ap.add_argument("xlsx", nargs="?")
    ap.add_argument("--self-test", action="store_true",
                    help="造几份已知有病的表，确认闸门还抓得住")
    ap.add_argument("--json", action="store_true")
    ap.add_argument("--freeze-min-rows", type=int, default=TABLE_MIN_ROWS,
                    help=f"超过多少行才要求冻结表头，默认 {TABLE_MIN_ROWS}")
    a = ap.parse_args()

    if a.self_test:
        return self_test()
    if not a.xlsx:
        ap.error("需要给一个 .xlsx 路径，或用 --self-test")

    try:
        import openpyxl                                   # noqa: F401
    except ImportError:
        print("⚠ 未安装 openpyxl，无法验表。装：pip install openpyxl", file=sys.stderr)
        return 2

    path = Path(a.xlsx)
    if not path.exists():
        print(f"找不到文件：{path}", file=sys.stderr)
        return 2

    try:
        fails, warns, notes, infos = verify(path, a.freeze_min_rows)
    except zipfile.BadZipFile:
        print("❌ FAIL：这不是一个能打开的 xlsx 包（zip 结构已损坏）")
        return 1

    fl = collapse(fails)
    wl = collapse(warns)

    if a.json:
        print(json.dumps({"fails": fl, "warns": wl, "skipped": notes,
                          "sheets": infos}, ensure_ascii=False, indent=2))
        return 1 if fl else 0

    line = "─" * 68
    print(line)
    print(f"Excel 自检：{path.name}   {len(infos)} 个 sheet")
    print(line)
    if fl:
        print(f"\n❌ FAIL {len(fl)} 项（必须修）")
        for x in fl:
            print(f"   - {x}")
    if wl:
        print(f"\n⚠ WARN {len(wl)} 项（自己判断）")
        for x in wl:
            print(f"   - {x}")
    if notes:
        print(f"\n○ 跳过 {len(notes)} 项（**不等于通过**）")
        for x in notes:
            print(f"   - {x}")
    if not fl and not wl:
        print("\n✅ 机器能查的都通过了。")

    print("\n各 sheet 概览")
    for i in infos:
        szs = "/".join(f"{s:g}" for s in i["font_sizes"][:8]) or "默认"
        print(f"   · 「{i['sheet']}」{i['rows']} 行 × {i['cols']} 列｜"
              f"合并 {i['merged']} 处｜冻结 {i['freeze'] or '无'}｜"
              f"显式列宽 {i['explicit_widths']} 列｜"
              f"字号 {szs}｜字色 {i['font_colors']} 种｜填充 {i['fills']} 种")

    print("\n机器只判「排版会不会出错」。这几件它判不了，自己打开表看一眼：")
    print("  · 数字的口径和单位在表里说清楚了没有（万元还是元、含不含税）")
    print("  · 该用公式的地方是不是硬编码成了值——用户改了输入，结果不跟着变")
    print("  · 图表引用的是「数据」sheet 还是写死的数组")
    print(line)
    return 1 if fl else 0


if __name__ == "__main__":
    sys.exit(main())
